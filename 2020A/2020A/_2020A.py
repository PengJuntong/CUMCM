import numpy as np
import xlrd
from matplotlib import pyplot as plt
import matplotlib
import pandas as pd
from function import *
import openpyxl

data = xlrd.open_workbook('data.xls')
table = data.sheet_by_name(u'Sheet1')

data_org = np.array(table.col_values(1)[1:])
time_org = np.array(table.col_values(0)[1:])
Delta_t = 0.5

#初始的温区设定
gy1 = [175,195,235,255]

#找最优k
def search_k(start, end, k_predict):
    speed = 70/60.0
    k_best = k_predict
    flag = False

    for k in np.arange(k_predict - 0.006 , k_predict + 0.008 , 0.0001):
        T_now = data_org[start]
        error_tmp = 0
        for t in range(start, end):
            T_next = T_now *(1 - k*Delta_t) + k * T_air(time_org[t]*speed,gy1) * Delta_t
            error_tmp += (T_next - data_org[t+1])**2
            T_now = T_next
        if flag:
            if(error_tmp < error):
               error = error_tmp
               k_best = k
        else: 
            error = error_tmp
            flag = True

    return [k_best, error]

k1 = search_k(0,291,0.017)
k2 = search_k(292,351,0.018)
k3 = search_k(352,412,0.025)
k4 = search_k(413,534,0.021)
k5 = search_k(534,708,0.02)
kerror = [k1,k2,k3,k4,k5]
#print(kerror)

cal_klist = [k1[0], k2[0], k3[0], k4[0], k5[0]] #每段最优k


#写入k值及误差
data_df=pd.DataFrame(kerror)
data_df.columns =["k","error"]
data_df.index = ["k1","k2","k3","k4","k5"]
writer = pd.ExcelWriter('..\..\klist.xls')  
data_df.to_excel(writer,float_format='%.7f')  
writer.save()

#根据初值生成炉温曲线
def u_cacl(speed,gy_x):
    T_cacl = [30.00]
    i = 0
    for t in np.arange(19.0,373.0, 0.5):
        distance = t*speed
        k = k_s(distance)
        T_next = T_cacl[i] *(1-k*Delta_t) + k * T_air(distance,gy_x) * Delta_t
        T_cacl.append(T_next)
        i += 1
    return T_cacl

#画图
def painter(y, title,filetitle,contrast):
    plt.figure()
    plt.rcParams['font.family']=['Microsoft Yahei']
    plt.title("%s" %title)
    plt.xlabel("时间(s)")
    plt.ylabel("温度(℃)")
    plt.plot(time_org,y)
    if contrast:
        plt.plot(time_org,data_org)
    plt.savefig('%s.png' %filetitle)
    #plt.show()

painter(u_cacl(70/60.0, gy1),
        '生成炉温曲线与原始炉温曲线的比较','prob1-1',True)

gy1_1 = [173,198,230,257]

answer_p1 = u_cacl(78/60.0,gy1_1)

painter(answer_p1,
        '新炉温曲线与原始炉温曲线的比较','prob1-2',True)



#写入答案
ans = [time_org,answer_p1]
ans = np.transpose(ans)
np.savetxt("result.csv",ans, delimiter = ',')



###p2
#更新小温区温度
gy2 = [182,203,237,254]

#斜率检验函数
def properrate(y):
    flag=True
    for i in range (len(y)-1):
        rate=y[i+1]-y[i]
        if abs(rate) > 1.5:
            flag=False
            break
        else:
            continue
    return flag

#峰值温度检验函数
def propermax(y):
    if max(y) > 240 and max(y) < 250:
        return True
    else:
        return False

#找高于217时间点
def highT_time(y):

    init_i=0
    final_i=len(y)-1

    for i in range (len(y)):
        if y[i]>217:
            init_i=i
            break
        else:
            continue
    for i in range(len(y)-1, 0, -1): #range(,,-1)反向遍历
        if y[i]>217:
            final_i = i
            break
        else :
            continue

    return [init_i, final_i]


#高于217的时间检验函数
def hightemtime(y):
    i = highT_time(y)
    deltai = i[1] - i[0]

    if(deltai>80 and deltai<180):
        return True
    else:
        return False

#温度区间持续时间检验函数
def sptime(y):

    init_i=0
    final_i=len(y)-1

    for i in range (len(y)):
        if y[i]>150:
            init_i=i
            break
        else:
            continue
    for i in range (init_i,len(y)-1):
        if y[i]>190:
            final_i=i
            break
        else:
            continue
    deltai = final_i - init_i

    if(deltai>120 and deltai<240):
        return True
    else:
        return False


#递归用生成函数
def u_cacl2(v):
    return u_cacl(v,gy2)

lowerbound = 0.5
upperbound = 4
 
step=0.1

def recurf1(x):
    global step
    if step<pow(0.1,5):
        return x

    temp=u_cacl2(x)
    if not propermax(temp):
        x=x-step
        return recurf1(x)
    if propermax(temp):
        x=x+step
        step=step/2.0
        return recurf1(x)

def recurf(x, algorithm):
    global step
    global upperbound
    if algorithm(u_cacl2(upperbound)): return upperbound
    if step<pow(0.1,5):
        return x

    temp=u_cacl2(x)
    if not algorithm(temp):
        x=x-step
        return recurf(x)
    if algorithm(temp):
        x=x+step
        step=step/2.0
        return recurf(x)

upperbound=recurf1(upperbound)
step=0.1
upperbound=recurf(upperbound, properrate)
step=0.1
upperbound=recurf(upperbound, hightemtime)
step=0.1
upperbound=recurf(upperbound, sptime)


print("问题2中的速度最大值为：")
print(upperbound)
np.savetxt("prob2.csv",[upperbound], delimiter = ',')
answer_p2 = u_cacl2(upperbound)

painter(answer_p2,
        '问题2炉温曲线','prob2',False)


###p3、p4

def chao(M,a,b):

    m0 = np.random.rand()
    m_list = [((b-a)*m0+a)]
    for i in range(M-1):
        m = 4*m0*(1-m0)
        m_list.append(((b-a)*m+a))
        m0 = m
    return m_list


#判断温度序列是否满足制程界限
def goodSetting(u_list):
    if properrate(u_list):
        if sptime(u_list):
            if propermax(u_list):
                if hightemtime:
                    return True
    else:
        return False

#算面积
def highT_area(u_list):

    max_time = u_list.index(max(u_list))
    i = highT_time(u_list)[0]
    area = 0
    for time in range(i,max_time):
        area += u_list[time]*Delta_t

    return area-((max_time-i)*u_list[i]*Delta_t)

#算镜像误差
def mirror_error(u_list):
    max_time = u_list.index(max(u_list))
    i = highT_time(u_list)[1]
    error = 0
    for j in range(max_time,i):
        error += abs(u_list[2*max_time-j]-u_list[j])
    return error

nrows=0
cycletimes = 10000
while True:
    tempre_pre = chao(cycletimes,165,185)
    tempre_stable1 = chao(cycletimes,185,205)
    tempre_stable2 = chao(cycletimes,225,245)
    tempre_back = chao(cycletimes,245,265)
    v_search = chao(cycletimes,65/60.0,100/60.0)
    P_mark = [False]*cycletimes #标记该设定是否满足制程界限

    #将5个混沌序列并排后转置，变成一行一组设定
    data_list = np.array([v_search, P_mark, tempre_pre, tempre_stable1, tempre_stable2, tempre_back])
    data_list = data_list.transpose()

    flag = False
    S_min = 0
    S_max = 0
    error_max = 0
    error_min = 0
    answer_p3 = np.zeros(5)

    #遍历求最小面积，顺便求最大面积和误差
    for u in data_list:
        u_list = u_cacl(u[0],u[2:])

        if goodSetting(u_list):
            u[1] = True #将符合制程界限的设定标记变成True

            S = highT_area(u_list) 
            error = mirror_error(u_list)
            #print(S,error)

            if flag:                
                if S <= S_min:
                    S_min = S
                    answer_p3 = u
                    
                if S >= S_max:
                    S_max = S
                if error <= error_min:
                    error_min = error
                if error >= error_max:
                    error_max = error
            else:
                S_min = S
                S_max = S
                error_min = error
                error_max = error
                flag = True
            wb=openpyxl.load_workbook("prob3.xlsx")
            temp=wb['Sheet1']
            temp=wb.active
            nrows = temp.max_row
            u=np.append(u,S)
            for j in range (6):
                temp.cell(row=nrows+1,column=j+1,value=u[j])
            wb.save("prob3.xlsx")
            print(u)
    
    flag = False
    P_best = 0
    answer_p4 = np.zeros(5)

    #遍历求综合指标
    for u in data_list:
        if u[1]:#只对标记为True的遍历，不再重复调用判断函数

            u_list = u_cacl(u[0],u[2:])
            S = highT_area(u_list) 
            error = mirror_error(u_list)
            P = (S-S_min)/(S_max-S_min) + (error-error_min)/(error_max-error_min)
            #print(P)

            if flag:                
                if P <= P_best:
                    P_best = P
                    answer_p4 = u
                    
            else:
                P_best = P
                flag = True
            wb=openpyxl.load_workbook("prob4.xlsx")
            temp=wb['Sheet1']
            temp=wb.active
            nrows = temp.max_row
            u=np.append(u,P)
            for j in range (6):
                temp.cell(row=nrows+1,column=j+1,value=u[j])
            wb.save("prob4.xlsx")
            print(u)

    

    
    